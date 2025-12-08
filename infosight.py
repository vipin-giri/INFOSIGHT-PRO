import argparse
import requests
import re
import sys
import socket
import time
import json
import os
import csv
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
from collections import deque
from datetime import datetime
import urllib3

SCAN_RESULTS = []

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ==========================
# Regex patterns
# ==========================
VERSION_RE = re.compile(r"\b\d+(?:\.\d+){0,}\b")
PROD_SLASH_VER_RE = re.compile(r"([A-Za-z0-9\-]+)/(\d+[A-Za-z0-9\.\-_]*)")
LEADING_V_RE = re.compile(r"\bv\d+(?:\.\d+)*\b", re.IGNORECASE)

# JWT-like tokens
JWT_RE = re.compile(
    r"\beyJ[A-Za-z0-9_\-]{10,}\.[A-Za-z0-9_\-]{10,}\.[A-Za-z0-9_\-]{10,}\b"
)

# Private IP ranges
PRIVATE_IP_RE = re.compile(
    r"\b(?:10\.\d{1,3}\.\d{1,3}\.\d{1,3}|"
    r"192\.168\.\d{1,3}\.\d{1,3}|"
    r"172\.(?:1[6-9]|2\d|3[0-1])\.\d{1,3}\.\d{1,3})\b"
)

# Email addresses
EMAIL_RE = re.compile(
    r"[a-zA-Z0-9_.+\-]+@[a-zA-Z0-9\-.]+\.[a-zA-Z0-9\-.]+"
)

# Cloud storage URLs
S3_URL_RE = re.compile(r"https?://[A-Za-z0-9\-\.]+\.s3\.amazonaws\.com[^\s\"'<>]*")
GCS_URL_RE = re.compile(r"https?://storage.googleapis.com/[^\s\"'<>]*")
AZURE_BLOB_RE = re.compile(r"https?://[A-Za-z0-9\-\.]+\.blob\.core\.windows\.net[^\s\"'<>]*")

# Common API key patterns (AWS, Google, Slack etc.)
AWS_KEY_RE = re.compile(r"\bAKIA[0-9A-Z]{16}\b")
GOOGLE_API_KEY_RE = re.compile(r"\bAIza[0-9A-Za-z\-_]{35}\b")
SLACK_TOKEN_RE = re.compile(r"\bxox[baprs]-[0-9A-Za-z\-]{10,48}\b")

JS_SECRET_RE = re.compile(
    r"(api_key|apikey|apiKey|secret|token)\s*[:=]\s*['\"][^'\"\\]{10,}['\"]",
    re.IGNORECASE
)

# ==========================
# Headers to inspect
# ==========================
DEFAULT_HEADERS = [
    "Server", "X-Powered-By", "X-AspNet-Version", "X-AspNetMvc-Version",
    "X-Powered-By-Plesk", "X-Powered-By-PHP", "X-PHP-Version", "X-Generator",
    "X-Drupal-Cache", "X-Drupal-Cache-Context", "X-Backend-Server", "X-Backend",
    "X-App-Server", "X-Service", "X-Instance-Id", "X-Instance", "X-Served-By",
    "X-Cache", "X-Cache-Lookup", "Via", "X-Varnish", "X-Amz-Cf-Id",
    "X-AspNetMvc-Version", "X-Pingback", "X-Redirect-By", "X-Turbo-Charged-By",
    "X-Debug-Token", "X-Debug-Token-Link", "X-Runtime", "X-Env",
    "X-Revision", "X-Version", "X-Build", "X-Commit", "X-Hostname"
]

# ==========================
# Color codes for terminal
# ==========================
class Colors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

# ==========================
# Mitigation tips & refs
# ==========================
MITIGATION_TIPS = [
    "Remove or obfuscate version numbers in Server and X-Powered-By headers",
    "Use reverse proxy (Nginx/Apache) to strip upstream headers",
    "Configure ServerTokens Prod (Apache) or server_tokens off (Nginx)",
    "Set expose_php = Off in php.ini",
    "Remove generator meta tags from HTML",
    "Sanitize HTML comments in production",
    "Restrict robots.txt to non-sensitive paths only",
    "Review sitemap.xml for internal/staging URLs",
    "Disable directory listing and server status pages",
    "Remove default error pages that reveal version info",
    "Avoid exposing API keys/tokens in JS, HTML, or hidden fields",
    "Avoid exposing internal IPs, environment info, and debug headers",
    "Use separate credentials for frontend-only public APIs"
]

REFERENCES = [
    "https://owasp.org/www-project-secure-headers/",
    "https://cheatsheetseries.owasp.org/cheatsheets/HTTP_Headers_Cheat_Sheet.html",
    "https://developer.mozilla.org/en-US/docs/Web/HTTP/Headers/Server",
    "https://owasp.org/www-community/attacks/Information_disclosure"
]

# ==========================
# Detection helpers
# ==========================
def has_version_info(value: str):
    """Return (is_version_present, evidence_string, product, version)"""
    if not value:
        return False, "", None, None

    m = PROD_SLASH_VER_RE.search(value)
    if m:
        return True, f"{m.group(1)}/{m.group(2)}", m.group(1), m.group(2)

    m2 = LEADING_V_RE.search(value)
    if m2:
        return True, m2.group(0), None, m2.group(0)

    m3 = VERSION_RE.search(value)
    if m3:
        return True, m3.group(0), None, m3.group(0)

    return False, "", None, None

def normalize_url(url):
    """Ensure URL has proper scheme"""
    if not url.startswith("http://") and not url.startswith("https://"):
        url = "https://" + url
    return url

def get_base_domain(url):
    """Extract base domain from URL"""
    parsed = urlparse(url)
    return f"{parsed.scheme}://{parsed.netloc}"

def is_same_domain(url1, url2):
    """Check if two URLs belong to the same domain"""
    return urlparse(url1).netloc == urlparse(url2).netloc

def severity_for(header_name: str, has_version: bool):
    """Determine severity level for header leaks"""
    h = header_name.lower()
    if has_version:
        if h == "server":
            return "High"
        if "x-powered" in h or h.startswith("x-aspnet") or "php" in h:
            return "Medium"
    return "Low"

# ==========================
# Crawl website for URLs
# ==========================
def crawl_website(base_url, verify_ssl=True, max_pages=None, max_depth=None, verbose=True):
    """
    Crawl website and return list of unique URLs to scan.
    """
    start_url = normalize_url(base_url)
    visited = set()
    to_visit = deque([(start_url, 0)])
    found_urls = []

    if verbose:
        print(f"\n{Colors.OKCYAN}[*] Starting crawl of {start_url} "
              f"(max_pages: {max_pages if max_pages is not None else 'UNLIMITED'}, "
              f"max_depth: {max_depth if max_depth is not None else 'UNLIMITED'})...{Colors.ENDC}")
        print(f"{Colors.OKCYAN}[*] Crawling in progress...{Colors.ENDC}\n")

    while to_visit:
        current_url, depth = to_visit.popleft()

        if current_url in visited:
            continue

        if max_pages is not None and len(visited) >= max_pages:
            break

        visited.add(current_url)
        found_urls.append(current_url)

        try:
            if verbose:
                print(f"  {Colors.OKBLUE}→{Colors.ENDC} Crawling: {current_url}")

            response = requests.get(current_url, timeout=10, verify=verify_ssl, allow_redirects=True)

            if 'text/html' in response.headers.get('Content-Type', ''):
                if max_depth is not None and depth >= max_depth:
                    continue

                soup = BeautifulSoup(response.text, 'html.parser')

                for link in soup.find_all('a', href=True):
                    href = link['href']
                    full_url = urljoin(current_url, href)

                    if not is_same_domain(full_url, base_url):
                        continue

                    full_url = full_url.split('#')[0].strip()

                    if full_url and full_url not in visited:
                        already_queued = any(u == full_url for (u, _) in to_visit)
                        if not already_queued:
                            to_visit.append((full_url, depth + 1))

        except Exception:
            if verbose:
                print(f"  {Colors.FAIL}✗{Colors.ENDC} Failed to crawl: {current_url}")

    if verbose:
        print(f"\n{Colors.OKGREEN}[*] Crawl completed. Found {len(found_urls)} unique pages to scan.{Colors.ENDC}\n")

    return found_urls

# ==========================
# Deep scanning functions
# ==========================
def check_sensitive_endpoints(url, verify_ssl=True):
    """Check for exposed sensitive endpoints"""
    findings = []
    endpoints = [
        "/server-status", "/status", "/phpinfo.php", "/.git/config",
        "/wp-login.php", "/admin", "/.env", "/.git/HEAD",
        "/composer.json", "/package.json", "/.htaccess", "/web.config",
        "/.svn/entries", "/backup.sql", "/backup.zip", "/db.sql",
        "/config.php", "/configuration.php", "/.DS_Store", "/crossdomain.xml"
    ]

    for ep in endpoints:
        try:
            test_url = url.rstrip("/") + ep
            resp = requests.head(test_url, timeout=6, verify=verify_ssl, allow_redirects=True)
            if resp.status_code and resp.status_code < 400:
                findings.append({
                    "type": "sensitive_endpoint",
                    "endpoint": ep,
                    "status_code": resp.status_code,
                    "severity": "High" if ep in ["/.env", "/.git/config", "/backup.sql", "/db.sql"] else "Medium"
                })
        except Exception:
            pass

    return findings

def analyze_html_deep(response, url, verify_ssl=True):
    """Deep HTML analysis for information disclosure"""
    findings = []

    try:
        body = response.text or ""
        ctype = response.headers.get("Content-Type", "")

        if not body or ("html" not in ctype.lower() and "<html" not in body.lower()):
            return findings

        soup = BeautifulSoup(body, "html.parser")

        # Meta generator
        meta = soup.find("meta", attrs={"name": re.compile(r"generator", re.I)})
        if meta and meta.get("content"):
            val = meta["content"]
            is_ver, evidence, prod, ver = has_version_info(val)
            findings.append({
                "type": "meta_generator",
                "value": val,
                "version_present": is_ver,
                "evidence": evidence,
                "product": prod,
                "version": ver,
                "severity": "Medium" if is_ver else "Low"
            })

        # HTML comments
        comments = re.findall(r"", body)
        interest_keywords = [
            "admin", "login", "wp-admin", "backup", ".env", "api_key",
            "api-key", "secret", "password", "passwd", "sql error",
            "exception", "traceback", "ora-", "mysql", "jdbc", "token",
            "debug", "TODO", "FIXME", "username", "credentials"
        ]

        for c in comments:
            lowered = c.lower()
            is_interesting = any(k in lowered for k in interest_keywords)

            m = PROD_SLASH_VER_RE.search(c) or LEADING_V_RE.search(c) or VERSION_RE.search(c)
            evidence = None
            if m:
                evidence = m.group(0)
                if re.fullmatch(r"\d", evidence):
                    evidence = None
                if evidence and not ('.' in evidence or evidence.startswith('v') or '/' in evidence):
                    evidence = None

            if not (is_interesting or evidence):
                continue

            sev = "Medium" if is_interesting else "Low"
            if is_interesting and any(k in lowered for k in ['api_key', '.env', 'passwd', 'password', 'secret', 'token']):
                sev = "High"

            findings.append({
                "type": "html_comment",
                "value": c.strip()[:500],
                "version_present": bool(evidence),
                "evidence": evidence,
                "severity": sev
            })
            break

        # Error messages
        error_patterns = [
            r"Fatal error:.*in.*on line \d+",
            r"Warning:.*in.*on line \d+",
            r"Parse error:.*in.*on line \d+",
            r"Traceback \(most recent call last\):",
            r"Exception in thread",
            r"ORA-\d{5}",
            r"MySQL Error",
            r"PSQLException"
        ]

        for pattern in error_patterns:
            matches = re.findall(pattern, body, re.IGNORECASE)
            if matches:
                findings.append({
                    "type": "error_message",
                    "value": matches[0][:300],
                    "severity": "High"
                })
                break

        # robots.txt
        try:
            robots_url = get_base_domain(url) + "/robots.txt"
            r2 = requests.get(robots_url, timeout=6, verify=verify_ssl)
            if r2.status_code == 200 and r2.text:
                lines = [l.strip() for l in r2.text.splitlines() if l.strip()]
                sensitive = [l for l in lines if re.search(
                    r"(admin|config|wp-admin|\.env|backup|backup-db|private|secret|api|dashboard)",
                    l, re.I
                )]
                if sensitive:
                    findings.append({
                        "type": "robots_txt",
                        "value": "\n".join(sensitive[:8]),
                        "severity": "Medium"
                    })
        except Exception:
            pass

        # sitemap.xml
        try:
            sitemap_url = get_base_domain(url) + "/sitemap.xml"
            r3 = requests.get(sitemap_url, timeout=6, verify=verify_ssl)
            if r3.status_code == 200 and r3.text:
                m = re.search(
                    r"(http[s]?://[\w\-.]*internal|http[s]?://[\w\-.]*staging|http[s]?://[\w\-.]*dev|http[s]?://[\w\-.]*test)",
                    r3.text, re.I
                )
                if m:
                    findings.append({
                        "type": "sitemap_leak",
                        "value": m.group(0),
                        "severity": "Medium"
                    })
        except Exception:
            pass

        # Directory listing
        if ("Index of /" in body) or re.search(r"<h1>Index of /", body, re.I):
            findings.append({
                "type": "directory_listing",
                "value": "Potential directory listing detected (Index of /)",
                "severity": "High"
            })

        # Emails
        emails = list(set(EMAIL_RE.findall(body)))
        if emails:
            findings.append({
                "type": "email",
                "value": ", ".join(emails[:10]),
                "severity": "Low"
            })

        # Private IPs
        ips = list(set(PRIVATE_IP_RE.findall(body)))
        if ips:
            findings.append({
                "type": "private_ip",
                "value": ", ".join(ips[:10]),
                "severity": "Medium"
            })

        # JWT tokens
        jwts = list(set(JWT_RE.findall(body)))
        if jwts:
            findings.append({
                "type": "jwt",
                "value": "; ".join(jwts[:3]),
                "severity": "Medium"
            })

        # Cloud storage URLs
        bucket_urls = set()
        bucket_urls.update(S3_URL_RE.findall(body))
        bucket_urls.update(GCS_URL_RE.findall(body))
        bucket_urls.update(AZURE_BLOB_RE.findall(body))
        if bucket_urls:
            findings.append({
                "type": "storage_bucket",
                "value": "\n".join(list(bucket_urls)[:10]),
                "severity": "Medium"
            })

        # API keys
        api_keys = set()
        api_keys.update(AWS_KEY_RE.findall(body))
        api_keys.update(GOOGLE_API_KEY_RE.findall(body))
        api_keys.update(SLACK_TOKEN_RE.findall(body))

        js_secret_matches = JS_SECRET_RE.findall(body)
        if js_secret_matches:
            api_keys.update([m[0] for m in js_secret_matches])

        if api_keys:
            findings.append({
                "type": "api_key",
                "value": ", ".join(list(api_keys)[:10]),
                "severity": "High"
            })

        # Hidden inputs
        hidden_inputs = soup.find_all("input", attrs={"type": "hidden"})
        for hi in hidden_inputs:
            name = (hi.get("name") or "").lower()
            value = (hi.get("value") or "")
            if not value:
                continue
            if any(k in name for k in ["token", "auth", "secret", "password", "passwd", "key"]):
                sev = "High" if any(k in name for k in ["secret", "password", "passwd"]) else "Medium"
                findings.append({
                    "type": "hidden_field",
                    "value": f"name={name}, value={value[:100]}",
                    "severity": sev
                })
                break

        # JS secrets
        js_secret_strings = re.findall(
            r"(api_key|apikey|apiKey|secret|token)\s*[:=]\s*['\"][^'\"\\]{10,}['\"]",
            body,
            flags=re.IGNORECASE
        )
        if js_secret_strings:
            findings.append({
                "type": "js_secret",
                "value": "; ".join(js_secret_strings[:5]),
                "severity": "High"
            })

    except Exception:
        pass

    return findings

# ==========================
# Scan single URL
# ==========================
def scan_single_url(url, verify_ssl=True, deep=True, active=False, show_url=True, verbose=True):
    """Check single URL for information disclosure."""
    findings = []

    # DNS check
    domain = url.replace("https://", "").replace("http://", "").split("/")[0]
    try:
        socket.gethostbyname(domain)
    except socket.error:
        if verbose and show_url:
            print(f"\n{Colors.WARNING}[⚠️] Skipping {url} — DNS not resolving.{Colors.ENDC}")
        return False, []

    response = None
    for attempt in range(2):
        try:
            response = requests.get(url, timeout=10, verify=verify_ssl, allow_redirects=True)
            break
        except requests.exceptions.SSLError:
            if verify_ssl and attempt == 0:
                if verbose and show_url:
                    print(f"{Colors.WARNING}[⚠️] SSL verification failed, retrying insecurely...{Colors.ENDC}")
                try:
                    response = requests.get(url, timeout=10, verify=False, allow_redirects=True)
                    break
                except Exception:
                    pass
        except requests.exceptions.RequestException as e:
            if verbose and show_url and attempt == 1:
                print(f"{Colors.FAIL}[ERROR] Request failed: {e}{Colors.ENDC}")
            elif attempt == 0:
                time.sleep(2)

    if not response:
        return False, []

    # Header version leaks
    for h in DEFAULT_HEADERS:
        val = response.headers.get(h)
        if val is None:
            continue

        is_ver, evidence, prod, ver = has_version_info(val)
        if is_ver:
            findings.append({
                "type": "header",
                "header": h,
                "value": val,
                "evidence": evidence,
                "product": prod,
                "version": ver,
                "severity": severity_for(h, is_ver)
            })

    # Deep scan
    if deep:
        deep_findings = analyze_html_deep(response, url, verify_ssl)
        findings.extend(deep_findings)

    # Active fingerprinting
    if active:
        endpoint_findings = check_sensitive_endpoints(url, verify_ssl)
        findings.extend(endpoint_findings)

    vulnerable = len(findings) > 0
    return vulnerable, findings

# ==========================
# Print findings to console
# ==========================
def print_findings(url, findings, show_mitigation=False):
    print(f"\n{Colors.BOLD}{'='*80}{Colors.ENDC}")
    print(f"{Colors.HEADER}{Colors.BOLD}TARGET: {url}{Colors.ENDC}")
    print(f"{Colors.BOLD}{'='*80}{Colors.ENDC}")

    if not findings:
        print(f"{Colors.OKGREEN}✓ No information disclosure vulnerabilities found{Colors.ENDC}")
        return

    high = [f for f in findings if f.get("severity") == "High"]
    medium = [f for f in findings if f.get("severity") == "Medium"]
    low = [f for f in findings if f.get("severity") == "Low"]

    print(f"\n{Colors.FAIL}{Colors.BOLD}VULNERABLE: Yes{Colors.ENDC}")
    print(f"Total Findings: {len(findings)} (High: {len(high)}, Medium: {len(medium)}, Low: {len(low)})\n")

    for f in findings:
        severity = f.get("severity", "Low")
        color = Colors.FAIL if severity == "High" else (Colors.WARNING if severity == "Medium" else Colors.OKBLUE)

        print(f"{color}[{severity.upper()}]{Colors.ENDC} ", end="")
        t = f.get("type")

        if t == "header":
            print(f"Header: {Colors.BOLD}{f['header']}{Colors.ENDC}")
            print(f"  Value: {f['value']}")
            if f.get("evidence"):
                print(f"  Evidence: {Colors.OKCYAN}{f['evidence']}{Colors.ENDC}")
        elif t == "meta_generator":
            print(f"Meta Generator: {f['value']}")
        elif t == "html_comment":
            print("HTML Comment (sensitive)")
            print(f"  Preview: {f['value'][:200]}...")
        elif t == "error_message":
            print("Error Message Exposed")
            print(f"  Message: {f['value'][:200]}...")
        elif t == "sensitive_endpoint":
            print(f"Sensitive Endpoint: {f['endpoint']} (Status: {f['status_code']})")
        elif t == "robots_txt":
            print("Robots.txt - Sensitive Paths Disclosed")
            print(f"  Paths: {f['value'][:200]}...")
        elif t == "sitemap_leak":
            print("Sitemap.xml - Internal URL Leak")
            print(f"  URL: {f['value']}")
        elif t == "directory_listing":
            print("Directory Listing Enabled")
            print(f"  Details: {f.get('value', '')[:200]}")
        else:
            label = t.replace("_", " ").title() if t else "Finding"
            print(label)
            if f.get("value"):
                print(f"  Details: {f['value'][:200]}...")
        print()

    if show_mitigation:
        print(f"\n{Colors.OKGREEN}{Colors.BOLD}MITIGATION RECOMMENDATIONS:{Colors.ENDC}")
        for tip in MITIGATION_TIPS:
            print(f"  • {tip}")

        print(f"\n{Colors.OKGREEN}{Colors.BOLD}REFERENCES:{Colors.ENDC}")
        for ref in REFERENCES:
            print(f"  • {ref}")

# ==========================
# Export helpers
# ==========================
def _aggregate_row_csv(r):
    """Row for CSV (simple summary)."""
    url = r['url']
    vuln = 'Yes' if r['vulnerable'] else 'No'
    finds = r['findings']
    ts = r['timestamp']

    total = len(finds)
    h = len([f for f in finds if f.get('severity') == 'High'])
    m = len([f for f in finds if f.get('severity') == 'Medium'])
    l = len([f for f in finds if f.get('severity') == 'Low'])

    types = sorted(set(f.get('type', '') for f in finds if f.get('type')))
    types_str = ', '.join(types)

    details = []
    evidences = set()
    for fnd in finds:
        t = fnd.get('type', '')
        if t == 'header':
            det = f"Header: {fnd.get('header')}={fnd.get('value')}"
        else:
            det = fnd.get('value', '')
        if det:
            details.append(det[:100])
        ev = fnd.get('evidence')
        if ev:
            evidences.add(str(ev))
    details_str = ' | '.join(details)[:300]
    evidence_str = ', '.join(sorted(evidences))[:200]

    # For CSV we still put counts in High/Med/Low
    return [
        url, vuln, total, h, m, l,
        types_str, details_str, evidence_str, ts
    ]

def _label_for_type(t):
    """Human-readable label for type."""
    mapping = {
        "header": "Server/Header",
        "meta_generator": "Meta generator",
        "html_comment": "HTML comment",
        "error_message": "Error message",
        "sensitive_endpoint": "Sensitive endpoint",
        "robots_txt": "robots.txt",
        "sitemap_leak": "sitemap.xml",
        "directory_listing": "Directory listing",
        "email": "Email",
        "private_ip": "Private IP",
        "jwt": "JWT token",
        "storage_bucket": "Cloud storage URL",
        "api_key": "API key",
        "hidden_field": "Hidden field",
        "js_secret": "JS secret",
    }
    return mapping.get(t, t.replace("_", " ").title() if t else "Finding")

def _aggregate_row_excel(r):
    """
    Row for Excel:
    URL, Vulnerable, Total, High✓, Med✓, Low✓, Indicators (multi-line), Evidence (multi-line), Time
    """
    url = r['url']
    vuln = 'Yes' if r['vulnerable'] else 'No'
    finds = r['findings']
    ts = r['timestamp']

    total = len(finds)
    h = len([f for f in finds if f.get('severity') == 'High'])
    m = len([f for f in finds if f.get('severity') == 'Medium'])
    l = len([f for f in finds if f.get('severity') == 'Low'])

    high_flag = '✓' if h > 0 else ''
    med_flag = '✓' if m > 0 else ''
    low_flag = '✓' if l > 0 else ''

    # group evidence by label
    label_to_vals = {}
    for fnd in finds:
        t = fnd.get('type', '')
        label = _label_for_type(t)
        if label not in label_to_vals:
            label_to_vals[label] = []
        if t == 'header':
            val = f"{fnd.get('header')}: {fnd.get('value')}"
        else:
            val = fnd.get('value') or ''
        ev = fnd.get('evidence')
        if ev and ev not in val:
            val = (val + " | " + str(ev)) if val else str(ev)
        if val:
            label_to_vals[label].append(val)

    indicators_lines = []
    evidence_lines = []
    for label, vals in label_to_vals.items():
        indicators_lines.append(label)
        # join multiple values for same label
        evidence_lines.append(" ; ".join(vals)[:200])

    indicators_str = "\n".join(indicators_lines)
    evidence_str = "\n".join(evidence_lines)

    return [
        url, vuln, total,
        high_flag, med_flag, low_flag,
        indicators_str, evidence_str, ts
    ]

# ==========================
# Exporters
# ==========================
def export_to_json(filename):
    try:
        filepath = os.path.join(os.getcwd(), filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(SCAN_RESULTS, f, indent=2, ensure_ascii=False)
        print(f"\n{Colors.OKGREEN}[✓] Results exported to: {filepath}{Colors.ENDC}")
    except Exception as e:
        print(f"\n{Colors.FAIL}[ERROR] Export failed: {e}{Colors.ENDC}")

def export_to_csv(filename):
    """CSV: one row per URL, counts, simple text."""
    try:
        filepath = os.path.join(os.getcwd(), filename)
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                'URL', 'Vulnerable', 'Total',
                'High', 'Med', 'Low',
                'Types', 'Details', 'Evidence', 'Time'
            ])
            for r in SCAN_RESULTS:
                writer.writerow(_aggregate_row_csv(r))
        print(f"\n{Colors.OKGREEN}[✓] Results exported to: {filepath}{Colors.ENDC}")
    except Exception as e:
        print(f"\n{Colors.FAIL}[ERROR] Export failed: {e}{Colors.ENDC}")

def export_to_excel(filename):
    """
    Excel: one row per URL, checkmarks for High/Med/Low,
    indicators + evidence in multi-line cells, with colours.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print(f"\n{Colors.WARNING}[!] Install openpyxl: pip install openpyxl{Colors.ENDC}")
        # fallback to CSV
        return export_to_csv(filename.replace('.xlsx', '.csv'))

    try:
        filepath = os.path.join(os.getcwd(), filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        headers = [
            'URL', 'Vulnerable', 'Total',
            'High', 'Med', 'Low',
            'Indicators', 'Evidence', 'Time'
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        row = 2
        for r in SCAN_RESULTS:
            agg = _aggregate_row_excel(r)
            for col, val in enumerate(agg, 1):
                cell = ws.cell(row, col, val)
                # wrap text for Indicators/Evidence
                if col in (7, 8):
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            # colour severity flags
            high_cell = ws.cell(row, 4)
            med_cell = ws.cell(row, 5)
            low_cell = ws.cell(row, 6)

            if high_cell.value == '✓':
                high_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                high_cell.font = Font(color="FFFFFF", bold=True)
            if med_cell.value == '✓':
                med_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                med_cell.font = Font(bold=True)
            if low_cell.value == '✓':
                low_cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
                low_cell.font = Font(color="FFFFFF", bold=True)

            row += 1

        widths = [45, 12, 8, 6, 6, 6, 30, 60, 22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        wb.save(filepath)
        print(f"\n{Colors.OKGREEN}[✓] Results exported to: {filepath}{Colors.ENDC}")
    except Exception as e:
        print(f"\n{Colors.FAIL}[ERROR] Export failed: {e}{Colors.ENDC}")

def export_to_txt(filename):
    """Plain text export, mostly for debugging."""
    try:
        filepath = os.path.join(os.getcwd(), filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("INFOSIGHT - SCAN RESULTS\n")
            f.write("=" * 80 + "\n\n")

            for r in SCAN_RESULTS:
                url, vuln, finds, ts = r['url'], r['vulnerable'], r['findings'], r['timestamp']
                f.write(f"{'='*80}\n")
                f.write(f"TARGET: {url}\n")
                f.write(f"TIMESTAMP: {ts}\n")
                f.write(f"{'='*80}\n\n")

                if not finds:
                    f.write("✓ No vulnerabilities found\n\n")
                else:
                    h = [x for x in finds if x.get('severity') == 'High']
                    m = [x for x in finds if x.get('severity') == 'Medium']
                    l = [x for x in finds if x.get('severity') == 'Low']

                    f.write(f"VULNERABLE: Yes\n")
                    f.write(f"Total: {len(finds)} (High: {len(h)}, Medium: {len(m)}, Low: {len(l)})\n\n")

                    for fi in finds:
                        sev = fi.get('severity', 'Low')
                        t = fi.get('type', 'finding')
                        val = fi.get('value', '')

                        f.write(f"[{sev.upper()}] {t}\n")
                        if t == 'header':
                            f.write(f"  Header: {fi.get('header')}\n")
                            f.write(f"  Value: {fi.get('value')}\n")
                            if fi.get('evidence'):
                                f.write(f"  Evidence: {fi.get('evidence')}\n")
                        else:
                            if val:
                                f.write(f"  Details: {val[:150]}...\n")
                        f.write("\n")

            f.write(f"{'='*80}\n")
            f.write(f"COMPLETED: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"{'='*80}\n")

        print(f"\n{Colors.OKGREEN}[✓] Results exported to: {filepath}{Colors.ENDC}")
    except Exception as e:
        print(f"\n{Colors.FAIL}[ERROR] Export failed: {e}{Colors.ENDC}")

# ==========================
# Scan target with crawling
# ==========================
def scan_target(url, verify_ssl=True, show_mitigation=False, only_vuln=False,
                crawl_mode=True, deep=True, active=False, max_pages=None, max_depth=None):
    """Scan target with optional crawling."""
    url = normalize_url(url)

    def record_result(u, vulnerable, findings):
        SCAN_RESULTS.append({
            'url': u,
            'vulnerable': vulnerable,
            'findings': findings,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })

    vulnerable, findings = scan_single_url(url, verify_ssl, deep, active, show_url=True, verbose=True)
    record_result(url, vulnerable, findings)

    if vulnerable or not only_vuln:
        print_findings(url, findings, show_mitigation)

    if not crawl_mode:
        return

    print(f"\n{Colors.OKCYAN}[*] Starting full website crawl...{Colors.ENDC}")
    base_url = get_base_domain(url)
    crawled_urls = crawl_website(base_url, verify_ssl, max_pages=max_pages, max_depth=max_depth)

    print(f"{Colors.OKCYAN}[*] Now scanning {len(crawled_urls)} discovered pages...{Colors.ENDC}\n")

    vuln_found = vulnerable
    for idx, crawled_url in enumerate(crawled_urls, 1):
        if crawled_url == url:
            continue

        print(f"[{idx}/{len(crawled_urls)}] Scanning: {crawled_url}")
        v, finds = scan_single_url(crawled_url, verify_ssl, deep, active, show_url=False, verbose=False)
        record_result(crawled_url, v, finds)

        if v:
            vuln_found = True
            if not only_vuln or (only_vuln and v):
                print_findings(crawled_url, finds, show_mitigation)

    if not vuln_found:
        print(f"\n{Colors.OKGREEN}[✓] Full website crawl completed. No vulnerabilities found.{Colors.ENDC}\n")

# ==========================
# Main
# ==========================
def main():

    print(Colors.FAIL + r"""
██╗███╗   ██╗███████╗ ██████╗ ███████╗██╗ ██████╗ ██╗  ██╗████████╗
██║████╗  ██║██╔════╝██╔═══██╗██╔════╝██║██╔════╝ ██║  ██║╚══██╔══╝
██║██╔██╗ ██║█████╗  ██║   ██║███████╗██║██║  ███╗███████║   ██║   
██║██║╚██╗██║██╔══╝  ██║   ██║╚════██║██║██║   ██║██╔══██║   ██║   
██║██║ ╚████║██║     ╚██████╔╝███████║██║╚██████╔╝██║  ██║   ██║   
╚═╝╚═╝  ╚═══╝╚═╝      ╚═════╝ ╚══════╝╚═╝ ╚═════╝ ╚═╝  ╚═╝   ╚═╝
""" + Colors.ENDC)

    parser = argparse.ArgumentParser(
        description="InfoSight - Advanced Information Disclosure Scanner with Web Crawling",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=""
    )

    parser.add_argument("-u", "--url", help="Target URL (e.g., https://example.com)")
    # Removed: parser.add_argument("target", nargs="?", help="Target URL (legacy positional, prefer -u/--url)")

    parser.add_argument("-f", "--file", help="File containing list of targets (one per line)")
    parser.add_argument("--no-verify", action="store_true", help="Skip SSL certificate verification")
    parser.add_argument("--show-mitigation", action="store_true", help="Show mitigation tips for vulnerabilities")
    parser.add_argument("--only-vuln", action="store_true", help="Show only vulnerable targets")

    parser.add_argument("--no-crawl", action="store_true", help="Disable web crawling (scan only the initial page)")

    parser.add_argument("--deep", action="store_true", default=True,
                        help="Deep scan (HTML, comments, robots.txt, etc.) - enabled by default")
    parser.add_argument("--no-deep", dest="deep", action="store_false", help="Disable deep scanning")
    parser.add_argument("--active", action="store_true", help="Active fingerprinting (test sensitive endpoints)")

    parser.add_argument("--max-pages", type=int, default=None,
                        help="Maximum pages to crawl per domain (default: unlimited)")
    parser.add_argument("--max-depth", type=int, default=None,
                        help="Maximum crawl depth (0=root only, 1=one hop, etc. default: unlimited)")

    parser.add_argument("--threads", type=int, default=5, help="Number of concurrent threads (default: 5)")
    parser.add_argument("-o", "--output", help="Export results to file (json/csv/xlsx/txt)")

    if len(sys.argv) == 1:
        print("Usage: infosight.py -u URL | -f FILE | [-h]")
        print("Run 'infosight.py -h' for full list of options.")
        print()
        parser.print_help()
        sys.exit(1)

    args = parser.parse_args()

    verify_ssl = not args.no_verify
    # Now relies only on args.url for single target mode
    single_target = args.url

    if not single_target and not args.file:
        print(f"{Colors.FAIL}[ERROR] You must provide either -u/--url or -f/--file.{Colors.ENDC}")
        parser.print_help()
        sys.exit(1)

    targets = []
    if args.file:
        try:
            with open(args.file, "r", encoding="utf-8") as fh:
                targets = [line.strip() for line in fh if line.strip()]
        except FileNotFoundError:
            print(f"{Colors.FAIL}[ERROR] File not found: {args.file}{Colors.ENDC}")
            sys.exit(1)
    elif single_target:
        targets = [single_target]

    print(f"\n{Colors.HEADER}{Colors.BOLD}{'='*80}{Colors.ENDC}")
    print(f"{Colors.HEADER}{Colors.BOLD}  INFOSIGHT - Advanced Information Disclosure Scanner{Colors.ENDC}")
    print(f"{Colors.HEADER}{Colors.BOLD}{'='*80}{Colors.ENDC}")
    print(f"\n{Colors.OKCYAN}Targets: {len(targets)}{Colors.ENDC}")
    print(f"{Colors.OKCYAN}Crawl Mode: {'ENABLED' if not args.no_crawl else 'DISABLED'}{Colors.ENDC}")
    print(f"{Colors.OKCYAN}Deep Scan: {'ENABLED' if args.deep else 'DISABLED'}{Colors.ENDC}")
    print(f"{Colors.OKCYAN}Active Mode: {'ENABLED' if args.active else 'DISABLED'}{Colors.ENDC}")
    print(f"{Colors.OKCYAN}SSL Verify: {'DISABLED' if args.no_verify else 'ENABLED'}{Colors.ENDC}\n")

    start_time = datetime.now()
    crawl_mode = not args.no_crawl

    if len(targets) == 1:
        scan_target(
            targets[0], verify_ssl, args.show_mitigation, args.only_vuln,
            crawl_mode, args.deep, args.active, args.max_pages, args.max_depth
        )
    else:
        with ThreadPoolExecutor(max_workers=args.threads) as executor:
            futures = [
                executor.submit(
                    scan_target, t, verify_ssl, args.show_mitigation,
                    args.only_vuln, crawl_mode, args.deep, args.active,
                    args.max_pages, args.max_depth
                )
                for t in targets
            ]
            for f in futures:
                f.result()

    if args.output:
        ext = args.output.lower().split('.')[-1]
        if ext == 'json':
            export_to_json(args.output)
        elif ext == 'csv':
            export_to_csv(args.output)
        elif ext in ['xlsx', 'xls']:
            export_to_excel(args.output)
        elif ext == 'txt':
            export_to_txt(args.output)
        else:
            print(f"{Colors.WARNING}[!] Unknown format, using TXT...{Colors.ENDC}")
            export_to_txt(args.output + '.txt')

    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()

    print(f"\n{Colors.OKGREEN}{Colors.BOLD}{'='*80}{Colors.ENDC}")
    print(f"{Colors.OKGREEN}{Colors.BOLD}[✓] Scan completed in {duration:.2f} seconds{Colors.ENDC}")
    print(f"{Colors.OKGREEN}{Colors.BOLD}{'='*80}{Colors.ENDC}\n")

if __name__ == "__main__":
    main()